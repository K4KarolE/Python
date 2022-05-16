# Movie Title Randomizer
# From my Excel we are pulling the "title", "year of release" and "how many times I have seen the movie" values
# The infos are in merged cells -> if the first randomly selected cell is empty, it is going to check the next cell until it finds a record*
#  *=after line 6736 there is still a title at the moment -> avoiding the search for valid cell on the empty/inactive part of the excel

import random
from openpyxl import Workbook, load_workbook
wb = load_workbook('D:/Movies.xlsx', data_only=True)
ws = wb.active

#data_only=True -> copying values from excel instead of formulas for the Haveseen cell value

cellnumber = random.randrange(6,6736)

cell = 'C' + str(cellnumber)
movietitle = ws[cell].value

cellRYear = 'E' + str(cellnumber)
ReleaseYear = ws[cellRYear].value


cellHSeen = 'N' + str(cellnumber)
HaveSeen = ws[cellHSeen].value


print()


if movietitle != None:
    print('Your movie this afternoon: ' + str(movietitle) + '(' + str(ReleaseYear) + ')')
    if HaveSeen == 1: 
        print('You have seen this movie only once since 05/2012.\n')
    else:
        print('You have seen this movie ' + str(HaveSeen) +  ' times since 05/2012.\n')
    quit() 

while movietitle == None:
    cellnumber += 1
    cell = 'C' + str(cellnumber)
    movietitle = ws[cell].value
    cellRYear = 'E' + str(cellnumber)
    ReleaseYear = ws[cellRYear].value
    cellHSeen = 'N' + str(cellnumber)
    HaveSeen = ws[cellHSeen].value
    
print('Your movie tonight: ' + str(movietitle) + '(' + str(ReleaseYear) + ')')

if HaveSeen == 1: 
    print('You have seen this movie only once since 05/2012.\n')
else:
    print('You have seen this movie ' + str(HaveSeen) +  ' times since 05/2012.\n')